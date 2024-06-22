import json
import pandas as pd
import numpy as np
import openpyxl
from sklearn.cluster import DBSCAN
from sklearn.preprocessing import StandardScaler

def juLeiFangFa(data):
    # 读取JSON文件
    with open('output.json', 'r', encoding='utf-8') as file:
        data = json.load(file)

    # 提取经度和纬度列数据
    coordinates = [(item['longitude'], item['latitude']) for item in data]
    coordinates = np.array(coordinates)

    # 特征缩放
    scaler = StandardScaler()
    scaled_coordinates = scaler.fit_transform(coordinates)

    # 运行DBSCAN算法
    dbscan = DBSCAN(eps=0.1, min_samples=2)
    dbscan.fit(scaled_coordinates)

    # 获取聚类标签
    labels = dbscan.labels_

    # 计算并输出分类数
    n_clusters = len(set(labels)) - (1 if -1 in labels else 0)
    print(f"分类数: {n_clusters}")

    # 输出每个聚类中心坐标
    unique_labels = np.unique(labels)
    cluster_centers = []
    for label in unique_labels:
        if label == -1:
            continue  # 跳过噪声点
        cluster_points = coordinates[labels == label]
        center = np.mean(cluster_points, axis=0)
        cluster_centers.append({
            "cluster_id": int(label + 1),
            "longitude": float(center[0]),
            "latitude": float(center[1])
        })

    # 以JSON格式返回聚类中心
    print(cluster_centers = [])

'''
# 创建一个空的字典，用于存储每个聚类中心的点
cluster_points = {label: [] for label in unique_labels if label != -1}

# 将每个数据点分配到相应的聚类中心
for i, label in enumerate(labels):
    if label != -1:
        cluster_points[label].append(coordinates[i])

# 输出每个聚类中心的所有点到Excel文件
wb = openpyxl.Workbook()
for label in cluster_points:
    cluster_data = np.array(cluster_points[label])
    sheet = wb.create_sheet(title='Cluster {}'.format(label + 1))
    for j, coord in enumerate(cluster_data):
        sheet.cell(row=j + 1, column=1).value = coord[0]
        sheet.cell(row=j + 1, column=2).value = coord[1]

# 保存Excel文件
wb.save('cluster_points.xlsx')'''

